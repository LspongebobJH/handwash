#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Sun Dec 11 12:56:51 2022

@author: cqy
"""
import cv2
import openpyxl
import os

raw_video_path = "/home/cqy/Documents/cqy/handwashing/hand_hygiene/rawdata/video/"
video_split_label_path = "/home/cqy/Documents/cqy/handwashing/hand_hygiene/rawdata/data_mark_COMP_1107.xlsx"
hand_hygiene_label_path = "/home/cqy/Documents/cqy/handwashing/hand_hygiene/rawdata/scanner_data_COMP_1107.xlsx"
segment_save_path = "/home/cqy/Documents/cqy/handwashing/hand_hygiene/dataset/"
segment_name_list = ["step_1","step_2","step_3","step_4","step_5","step_6","step_7","step_unknown"]

    
def read_video_split_label(video_name):
    wb = openpyxl.load_workbook(video_split_label_path)
    ws = wb.worksheets[0]
    row_num_list, row_value_list = value_by_col(ws, video_name, 1) #search for the video name in the first column
    
    for rep, row_num in enumerate(row_num_list):
        row_value = row_value_list[rep]
        for step in range(7):
            start_time = row_value[step*3]
            end_time = row_value[step*3+1]
            score = row_value[step*3+2]
            # print('start end score', start_time, end_time, score)
            if (score != 0 and score !=None):
                get_video_and_split(raw_video_path, video_name, step+1, rep, start_time, end_time, score)
                
        # step_1_stat_t = 
        # step_1_end_t =
        # step_1_move = 
        
        # step_2_stat_t = 
        # step_2_end_t =
        # step_2_move = 
        
        # step_3_stat_t = 
        # step_3_end_t =
        # step_3_move = 
        
        # step_4_stat_t = 
        # step_4_end_t =
        # step_4_move = 
        
        # step_5_stat_t = 
        # step_5_end_t =
        # step_5_move = 
        
        # step_6_stat_t = 
        # step_6_end_t =
        # step_6_move = 
        
        # step_7_stat_t = 
        # step_7_end_t =
        # step_7_move = 
        
        # split_mark = []
        
def value_by_col(ws, value, col):
    """
    :param value: 查询的值，需要与对应单元格数据类型一致。如查整数用int，文本用str，小数用flote
    :param col: 要查询的列号，从1开始，A对应1，B对应2，依次类推
    :return: NULL
    """
    max_row = ws.max_row
    max_col = ws.max_column
    # 针对指定列查询
    row_num_list = []
    row_value_list = []
    re = False
    for row in range(max_row):
        row_value = []
        use_cell = ws.cell(row + 1, col)
        if value == use_cell.value:
            print("已找到，在%s单元格"% use_cell.coordinate)
            row_num_list.append(row)
            row_value = [ws.cell(row + 1, col_num).value for col_num in range(4,25)]
            row_value_list.append(row_value)
            re = True
    if re:
        pass
    else:
        print("未找到符合的单元格")
    return row_num_list, row_value_list

def get_video_and_split(video_path, video_name, wash_step, rep, start_time, end_time, score):
    """
    

    Parameters
    ----------
    video_path : TYPE
        path for the saved raw videos.
    video_name : TYPE
        name of the video.
    wash_step : TYPE
        hand washing step to be extracted.
    rep : TYPE
        num of the repeated step.
    start_time : TYPE
        DESCRIPTION.
    end_time : TYPE
        DESCRIPTION.
    score : TYPE
        DESCRIPTION.

    Returns
    -------
    None.

    """
    cap  = cv2.VideoCapture(video_path+video_name)
    frames = cap.get(cv2.CAP_PROP_FRAME_COUNT)
    fps = cap.get(cv2.CAP_PROP_FPS)
    width = cap.get(cv2.CAP_PROP_FRAME_WIDTH)
    height = cap.get(cv2.CAP_PROP_FRAME_HEIGHT)
    
    #define the video format for the video segment
    fourcc = cv2.VideoWriter_fourcc(*'mp4v')
    segment_name_dir = segment_save_path+str(wash_step)+"/"+str(score)+"/"
    if not os.path.exists(segment_name_dir):
        os.makedirs(segment_name_dir)
    out = cv2.VideoWriter(segment_name_dir+ video_name.split('.')[0] + "_washstep"+ str(wash_step) + "_repeat" + str(rep) +".mp4", fourcc, fps, (int(width), int(height)))
    
    if (start_time!=None and end_time!=None and start_time!='n' and end_time!='n' ):
        #revise the start and end time
        
        start = (start_time-int(start_time))*100+int(start_time)*60
        end = (end_time-int(end_time))*100+int(end_time)*60
        
        start = round(start)
        end = round(end)
        
        if (start*fps <= frames and end*fps <= frames):
            # print("start and end", start, end)
            #cut the video
            cap.set(cv2.CAP_PROP_POS_FRAMES, start * fps)
            pos = cap.get(cv2.CAP_PROP_POS_FRAMES)  # get the start frame number
            while (pos <= end * fps ):
                
                ret, frame = cap.read()  # capture one frame
                out.write(frame)  # save the captured frame
                pos = cap.get(cv2.CAP_PROP_POS_FRAMES)
            out.release() 
            cap.release()
        else:
            print(video_name+'mark error in: start time '+ str(start) +' end time ' +str(end) )
        
if __name__ == "__main__":
    video_name_list = os.listdir(raw_video_path)
    for i, video_name in enumerate(video_name_list):
        read_video_split_label(video_name)